from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from manotron.db import init_db, session_scope
from manotron.export import export_to_excel
from manotron.models import InvoiceLine, LineLocationDeduction, SourceFile
from manotron.scan import MockInvoiceExtractor, scan_folder
from manotron.schemas import ExportOptions, ManotronSettings


def test_scan_skips_duplicate_hash_and_exports(tmp_path: Path) -> None:
    invoice_folder = tmp_path / "invoices"
    nested = invoice_folder / "nested"
    nested.mkdir(parents=True)
    sample = nested / "invoice.txt"
    sample.write_text("dummy invoice")

    db_path = tmp_path / "manotron.sqlite3"
    settings = ManotronSettings(
        watch_folder=str(invoice_folder),
        db_path=str(db_path),
        openai_api_key="test",
    )

    first = scan_folder(invoice_folder, settings, MockInvoiceExtractor())
    second = scan_folder(invoice_folder, settings, MockInvoiceExtractor())

    assert first.files_seen == 1
    assert first.files_processed == 1
    assert second.files_seen == 1
    assert second.files_skipped == 1

    init_db(db_path)
    with session_scope(db_path) as session:
        assert len(session.scalars(select(SourceFile)).all()) == 1
        assert len(session.scalars(select(InvoiceLine)).all()) == 1
        assert len(session.scalars(select(LineLocationDeduction)).all()) == 1

    output = export_to_excel(settings, ExportOptions(output_path=tmp_path / "out.xlsx"))
    workbook = load_workbook(output)
    sheet = workbook["invoice_lines"]
    assert sheet.max_row == 2
    assert sheet["A2"].value.startswith("MOCK-INVOICE")

